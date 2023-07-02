import os.path
from openpyxl import load_workbook
from path import dir_res


def test_read_xlsx_file():
    # Задаем уникальный путь
    xlsx_file = os.path.join(dir_res, 'file_example_XLSX_50.xlsx')
    # xlsx_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '../resources/file_example_XLSX_50.xlsx'))
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    # Проверяем результаты
    assert headers == [0, 'First Name', 'Last Name', 'Gender', 'Country', 'Age', 'Date', 'Id']

    # Проверяем результаты
    assert sheet.cell(row=2, column=2).value == 'Dulce'
